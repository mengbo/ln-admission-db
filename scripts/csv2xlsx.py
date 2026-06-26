#!/usr/bin/env python3
"""CSV → XLSX 转换脚本（用 openpyxl，走 uv 虚拟环境）。

用法：
    .venv/bin/python scripts/csv2xlsx.py input.csv output.xlsx
"""

import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def convert(csv_path: str, xlsx_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    with open(csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print(f"警告：{csv_path} 为空", file=sys.stderr)
        return

    # 写入数据
    for row_idx, row in enumerate(rows, 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center", horizontal="center")
            else:
                cell.alignment = wrap_alignment

    # 自动列宽（采样前 100 行，取最大宽度，封顶 50 字符）
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 100) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        col_width = min(max_len + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)
    print(f"✅ {csv_path} → {xlsx_path}（{len(rows) - 1} 行数据）")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(f"用法: {sys.argv[0]} input.csv output.xlsx", file=sys.stderr)
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
